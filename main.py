import sqlite3
import openpyxl
import os
import shutil
from datetime import datetime

DB_FILE = "musteri_kayit.db"

def db_baglanti():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""CREATE TABLE IF NOT EXISTS musteriler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ad_soyad TEXT NOT NULL,
        telefon TEXT,
        email TEXT,
        adres TEXT,
        kayit_zamani TEXT
    )""")
    return conn

def simdi():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def musteri_var_mi(conn, ad, tel):
    cur = conn.execute("SELECT 1 FROM musteriler WHERE ad_soyad=? AND telefon=?", (ad, tel))
    return cur.fetchone() is not None

def musteri_ekle(conn):
    ad = input("Ad Soyad: ")
    tel = input("Telefon: ")
    email = input("E-posta: ")
    adres = input("Adres: ")
    if not musteri_var_mi(conn, ad, tel):
        conn.execute("INSERT INTO musteriler (ad_soyad, telefon, email, adres, kayit_zamani) VALUES (?, ?, ?, ?, ?)",
                     (ad, tel, email, adres, simdi()))
        conn.commit()
        print("Müşteri eklendi.")
    else:
        print("Bu müşteri zaten kayıtlı!")

def musterileri_listele(conn):
    for row in conn.execute("SELECT * FROM musteriler ORDER BY id"):
        print(row)

def musteri_ara(conn):
    kelime = input("Arama: ")
    for row in conn.execute("""SELECT * FROM musteriler 
                                WHERE ad_soyad LIKE ? OR telefon LIKE ? OR email LIKE ? OR adres LIKE ?""",
                            (f"%{kelime}%",)*4):
        print(row)

def musteri_sil(conn):
    id_ = input("Silinecek ID: ")
    conn.execute("DELETE FROM musteriler WHERE id=?", (id_,))
    conn.commit()
    print("Müşteri silindi.")

def musteri_guncelle(conn):
    id_ = input("Güncellenecek ID: ")
    ad = input("Yeni Ad Soyad: ")
    tel = input("Yeni Telefon: ")
    email = input("Yeni E-posta: ")
    adres = input("Yeni Adres: ")
    conn.execute("""UPDATE musteriler 
                    SET ad_soyad=?, telefon=?, email=?, adres=? WHERE id=?""",
                 (ad, tel, email, adres, id_))
    conn.commit()
    print("Müşteri güncellendi.")

def excel_aktar_tumu(conn):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    ws.append(["ID", "Ad Soyad", "Telefon", "E-posta", "Adres", "Kayıt Zamanı"])
    for row in conn.execute("SELECT * FROM musteriler ORDER BY id"):
        ws.append(row)
    dosya_adi = "musteriler.xlsx"
    wb.save(dosya_adi)
    print(f"Tüm müşteriler Excel'e aktarıldı: {dosya_adi}")

def excel_aktar_tarih_araligi(conn):
    baslangic = input("Başlangıç tarihi (YYYY-MM-DD): ")
    bitis = input("Bitiş tarihi (YYYY-MM-DD): ")
    try:
        datetime.strptime(baslangic, "%Y-%m-%d")
        datetime.strptime(bitis, "%Y-%m-%d")
    except ValueError:
        print("Tarih formatı hatalı.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    ws.append(["ID", "Ad Soyad", "Telefon", "E-posta", "Adres", "Kayıt Zamanı"])
    query = """SELECT * FROM musteriler 
               WHERE DATE(kayit_zamani) BETWEEN ? AND ?
               ORDER BY kayit_zamani"""
    for row in conn.execute(query, (baslangic, bitis)):
        ws.append(row)

    dosya_adi = f"musteriler_{baslangic}_to_{bitis}.xlsx"
    wb.save(dosya_adi)
    print(f"{baslangic} - {bitis} arası müşteriler Excel'e aktarıldı: {dosya_adi}")

def excel_ice_aktar(conn):
    dosya_adi = input("İçe aktarılacak Excel dosya adı: ")
    if not os.path.exists(dosya_adi):
        print("Dosya bulunamadı.")
        return
    wb = openpyxl.load_workbook(dosya_adi)
    ws = wb.active
    satir_sayisi = 0
    atlanan = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        ad = row[1]
        tel = row[2]
        email = row[3]
        adres = row[4]
        tarih = row[5] if len(row) > 5 and row[5] else simdi()
        if ad and not musteri_var_mi(conn, ad, tel):
            conn.execute("INSERT INTO musteriler (ad_soyad, telefon, email, adres, kayit_zamani) VALUES (?, ?, ?, ?, ?)",
                         (ad, tel, email, adres, tarih))
            satir_sayisi += 1
        else:
            atlanan += 1
    conn.commit()
    print(f"{satir_sayisi} müşteri içe aktarıldı, {atlanan} kayıt atlandı (mükerrer).")

def son_eklenenler(conn):
    try:
        adet = int(input("Kaç son kayıt listelensin?: "))
    except ValueError:
        print("Geçersiz sayı!")
        return
    for row in conn.execute("SELECT * FROM musteriler ORDER BY kayit_zamani DESC LIMIT ?", (adet,)):
        print(row)

def otomatik_yedekle():
    if os.path.exists(DB_FILE):
        backup_name = f"musteri_kayit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        shutil.copy(DB_FILE, backup_name)
        print(f"Veritabanı yedeklendi: {backup_name}")

def menu():
    conn = db_baglanti()
    while True:
        print("""
1) Müşteri Ekle
2) Listele
3) Ara
4) Sil
5) Güncelle
6) Excel'e Aktar (Tüm)
7) Excel'e Aktar (Tarih Aralığı)
8) Excel'den İçe Aktar
9) Son Eklenen Müşteriler
0) Çıkış
""")
        secim = input("Seçim: ")
        if secim == "1": musteri_ekle(conn)
        elif secim == "2": musterileri_listele(conn)
        elif secim == "3": musteri_ara(conn)
        elif secim == "4": musteri_sil(conn)
        elif secim == "5": musteri_guncelle(conn)
        elif secim == "6": excel_aktar_tumu(conn)
        elif secim == "7": excel_aktar_tarih_araligi(conn)
        elif secim == "8": excel_ice_aktar(conn)
        elif secim == "9": son_eklenenler(conn)
        elif secim == "0":
            otomatik_yedekle()
            break
        else:
            print("Hatalı seçim.")

if __name__ == "__main__":
    menu()
